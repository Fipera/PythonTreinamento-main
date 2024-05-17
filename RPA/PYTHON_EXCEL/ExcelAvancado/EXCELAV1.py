#QUEBRANDO NOMES NA MESMA PLANILHA


from openpyxl import load_workbook
import os


nome_arquivo = 'C:\\Users\\Win7\\Desktop\\Documents\\RPA\\ExcelAvançado_Quebras_Email\\Quebrar.xlsx'
planilha_aberta = load_workbook(filename=nome_arquivo)


sheet_selecionada = planilha_aberta['Dados']

nomeNovo = ""
total_linha = len(sheet_selecionada['A']) + 1

for linha in range(2, total_linha):
    
    nomeAtual = sheet_selecionada['A%s' % linha].value

    if nomeNovo == nomeAtual:
        
        linhaSheetQuebra = len(sheet_selecionada2['A']) + 1
        colunaA = "A" +  str(linhaSheetQuebra)
        colunaB = "B" +  str(linhaSheetQuebra)
        colunaC = "C" +  str(linhaSheetQuebra)

        sheet_selecionada2[colunaA] = sheet_selecionada['A%s' % linha].value
        sheet_selecionada2[colunaB] = sheet_selecionada['B%s' % linha].value
        sheet_selecionada2[colunaC] = sheet_selecionada['C%s' % linha].value

    else:
        sheet_resumo = planilha_aberta.create_sheet(title=nomeAtual)

        sheet_selecionada2 = planilha_aberta[nomeAtual]

        nomeNovo = sheet_selecionada['A%s' % linha].value

        sheet_selecionada2['A1'] = 'Vendedor'
        sheet_selecionada2['B1'] = 'Produtos'
        sheet_selecionada2['C1'] = 'Vendas'

        sheet_selecionada2['A2'] = sheet_selecionada['A%s' % linha].value
        sheet_selecionada2['B2'] = sheet_selecionada['B%s' % linha].value
        sheet_selecionada2['C2'] = sheet_selecionada['C%s' % linha].value
    
planilha_aberta.save(filename=nome_arquivo)

os.startfile(nome_arquivo)