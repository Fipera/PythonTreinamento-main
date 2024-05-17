#QUEBRANDO NOMES PARA SEPARAR EM DIFERENTES PLANILHAS (SÓ QUE COLORIDO)

from openpyxl import load_workbook
from openpyxl import Workbook

from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.styles import colors
from openpyxl.cell import Cell

import os



nome_arquivo = 'C:\\Users\\Win7\\Desktop\\Documents\\RPA\\ExcelAvançado_Quebras_Email\\Quebrar.xlsx'
planilha_aberta = load_workbook(filename=nome_arquivo)


sheet_selecionada = planilha_aberta['Dados']

criandoNovoArquivoExcel = Workbook()

nomeNovo = ""
total_linha = len(sheet_selecionada['A']) + 1

corAzul = PatternFill(start_color='0099CCFF',
                     end_color='0099CCFF',
                     fill_type='solid')

corAmarelo = PatternFill(start_color='00FFFFCC',
                     end_color='00FFFFCC',
                     fill_type='solid')

bdFina= Side(border_style="thin", color="000000")
borda = Border(left=bdFina, right=bdFina, top=bdFina, bottom=bdFina)

for linha in range(2, total_linha):
    
    nomeAtual = sheet_selecionada['A%s' % linha].value

    if nomeNovo == nomeAtual:
        
        linhaSheetQuebra = len(selecionaSheetVendasNovaPlanilha['A']) + 1
        colunaA = "A" +  str(linhaSheetQuebra)
        colunaB = "B" +  str(linhaSheetQuebra)
        colunaC = "C" +  str(linhaSheetQuebra)

        selecionaSheetVendasNovaPlanilha[colunaA] = sheet_selecionada['A%s' % linha].value
        selecionaSheetVendasNovaPlanilha[colunaB] = sheet_selecionada['B%s' % linha].value
        selecionaSheetVendasNovaPlanilha[colunaC] = sheet_selecionada['C%s' % linha].value

      
        selecionaSheetVendasNovaPlanilha[colunaA].fill = corAmarelo
        selecionaSheetVendasNovaPlanilha[colunaB].fill = corAmarelo
        selecionaSheetVendasNovaPlanilha[colunaC].fill = corAmarelo

        criandoNovoArquivoExcel.save(filename=caminhoNovaPlanilha)

    else:

        nomeNovo = sheet_selecionada['A%s' % linha].value

        nova_planilha = criandoNovoArquivoExcel.active

        nova_planilha.title = "Vendas"

        caminhoNovaPlanilha = 'C:\\Users\\Win7\\Desktop\\Documents\\RPA\\ExcelAvançado_Quebras_Email\\' + sheet_selecionada['A%s' % linha].value + '.xlsx'

        selecionaSheetVendasNovaPlanilha = criandoNovoArquivoExcel['Vendas']

        selecionaSheetVendasNovaPlanilha['A1'] = 'Vendedor'
        selecionaSheetVendasNovaPlanilha['B1'] = 'Produtos'
        selecionaSheetVendasNovaPlanilha['C1'] = 'Vendas'

        selecionaSheetVendasNovaPlanilha['A2'] = sheet_selecionada['A%s' % linha].value
        selecionaSheetVendasNovaPlanilha['B2'] = sheet_selecionada['B%s' % linha].value
        selecionaSheetVendasNovaPlanilha['C2'] = sheet_selecionada['C%s' % linha].value
    
        selecionaSheetVendasNovaPlanilha['A1'].fill = corAzul
        selecionaSheetVendasNovaPlanilha['B1'].fill = corAzul
        selecionaSheetVendasNovaPlanilha['C1'].fill = corAzul
        selecionaSheetVendasNovaPlanilha['A1'].border = borda
        selecionaSheetVendasNovaPlanilha['B1'].border = borda
        selecionaSheetVendasNovaPlanilha['C1'].border = borda

        selecionaSheetVendasNovaPlanilha['A2'].fill = corAmarelo
        selecionaSheetVendasNovaPlanilha['B2'].fill = corAmarelo
        selecionaSheetVendasNovaPlanilha['C2'].fill = corAmarelo

        selecionaSheetVendasNovaPlanilha.column_dimensions["A"].width = 18
        selecionaSheetVendasNovaPlanilha.column_dimensions["B"].width = 25
        selecionaSheetVendasNovaPlanilha.column_dimensions["C"].width = 15
        selecionaSheetVendasNovaPlanilha.delete_rows(3, 100)

        criandoNovoArquivoExcel.save(filename=caminhoNovaPlanilha)

planilha_aberta.save(filename=nome_arquivo)

os.startfile(nome_arquivo)