import pandas as opcoesDoPanda
import os

caminhoArquivos = r"C:\Users\Win7\Desktop\Documents\RPA\Pandas\Excel"

listaArquivos = os.listdir(caminhoArquivos)


listaCaminhoArquivoExcel = [caminhoArquivos + "\\" + arquivo for arquivo in listaArquivos if arquivo[-4:] == 'xlsx' ]



dadosArquivo = opcoesDoPanda.DataFrame()

for arquivo in listaCaminhoArquivoExcel:
    dados = opcoesDoPanda.read_excel(arquivo)
    dadosArquivo = opcoesDoPanda.concat([dadosArquivo, dados])

dadosArquivo.to_excel(r"C:\Users\Win7\Desktop\Documents\RPA\Pandas\Arquivo Consolidado.xlsx")

#######################################

from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.styles import colors
from openpyxl.cell import Cell

caminhoArquivoDadosSistema = r"C:\Users\Win7\Desktop\Documents\RPA\Pandas\Arquivo Consolidado.xlsx"
planilhaDadosSistema = load_workbook(filename=caminhoArquivoDadosSistema)

sheet_selecionada = planilhaDadosSistema['Sheet1']

sheet_selecionada.delete_cols(1)

sheet_selecionada.title = "Dados Consolidados"

sheet_selecionada.column_dimensions['A'].width = 35
sheet_selecionada.column_dimensions['B'].width = 40


corCinza = PatternFill(start_color='00FFFFFF',
                     end_color = '00FFFFFF',
                     fill_type='solid'
)

corAmarela = PatternFill(start_color='00FFFFCC',
                     end_color = '00FFFFCC',
                     fill_type='solid'
)

BdFina = Side(border_style= 'thin', color='000000')
borda = Border(left=BdFina, right=BdFina, top=BdFina, bottom=BdFina)

sheet_selecionada['A1'].fill = corAmarela
sheet_selecionada['B1'].fill = corAmarela
sheet_selecionada['C1'].fill = corAmarela

sheet_selecionada['A1'].border = borda
sheet_selecionada['B1'].border = borda
sheet_selecionada['C1'].border = borda

for linha in range(2, len(sheet_selecionada['A']) + 1):

    celulaColunaA = 'A' + str(linha)
    celulaColunaB = 'B' + str(linha)
    celulaColunaC = 'C' + str(linha)

    sheet_selecionada[celulaColunaA].fill = corCinza
    sheet_selecionada[celulaColunaB].fill = corCinza
    sheet_selecionada[celulaColunaC].fill = corCinza

    sheet_selecionada[celulaColunaA].border = borda
    sheet_selecionada[celulaColunaB].border = borda
    sheet_selecionada[celulaColunaC].border = borda

FinalLinha = linha + 1

celulaFinalLinha = 'C' + str(FinalLinha)

formulaSoma = "=SUM(C2:C" + str(linha) + ")"

sheet_selecionada[celulaFinalLinha] = formulaSoma

planilhaDadosSistema.save(filename=caminhoArquivoDadosSistema)

os.startfile(caminhoArquivoDadosSistema)