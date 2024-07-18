from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
import os

excel_path = 'C:\\Users\\filip\\OneDrive\\Documentos\\RPA\\challenge.xlsx'
planilha_desafio = load_workbook(excel_path)
sheet_selecionada = planilha_desafio['Sheet1']

driver = webdriver.Firefox()


driver.get('https://rpachallenge.com/')
sleep(3)


for linha in sheet_selecionada.iter_rows(min_row=2, values_only=True):
    firstName, lastName, companyName, companyRole, address, email, number, *extra = linha


    if firstName and lastName and companyName and companyRole and address and email and number:




        adress_element = driver.find_element(By.CSS_SELECTOR, "[ng-reflect-name='labelAddress']")
        companyRole_element = driver.find_element(By.CSS_SELECTOR, "[ng-reflect-name='labelRole']")
        email_element = driver.find_element(By.CSS_SELECTOR, "[ng-reflect-name='labelEmail']")
        lastName_element = driver.find_element(By.CSS_SELECTOR, "[ng-reflect-name='labelLastName']")
        number_element = driver.find_element(By.CSS_SELECTOR, "[ng-reflect-name='labelPhone']")
        companyName_element = driver.find_element(By.CSS_SELECTOR, "[ng-reflect-name='labelCompanyName']")
        firstName_element = driver.find_element(By.CSS_SELECTOR, "[ng-reflect-name='labelFirstName']")
        sleep(1)

        adress_element.clear()
        adress_element.send_keys(address)
        
        companyRole_element.clear()
        companyRole_element.send_keys(companyRole)
        
        email_element.clear()
        email_element.send_keys(email)
        
        lastName_element.clear()
        lastName_element.send_keys(lastName)
        
        number_element.clear()
        number_element.send_keys(number)
        
        companyName_element.clear()
        companyName_element.send_keys(companyName)
        
        firstName_element.clear()
        firstName_element.send_keys(firstName)
        sleep(2)

        driver.find_element(By.XPATH, "/html/body/app-root/div[2]/app-rpa1/div/div[2]/form/input").click()
        




